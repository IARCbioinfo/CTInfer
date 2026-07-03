


import os, re, sys, time, threading, json, base64
from pathlib import Path

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import tkinter.scrolledtext as _st
except ImportError:
    print("ERREUR : tkinter non disponible."); sys.exit(1)


try:
    from translations import t, LANGUAGES, LANG_CODES, LANG_FLAGS, LANG_NAMES, DEFAULT_LANG, FLAG_IMAGES_B64
except ImportError:
    LANGUAGES = [("fr", "🇫🇷", "Français")]
    LANG_CODES = ["fr"]
    LANG_FLAGS = {"fr": "🇫🇷"}
    LANG_NAMES = {"fr": "Français"}
    DEFAULT_LANG = "fr"
    FLAG_IMAGES_B64 = {}
    def t(key, lang="fr", **kwargs): return key


def _auto_install(pkg, import_name=None):
    try:
        __import__(import_name or pkg)
    except ImportError:
        import subprocess as _sp, sys as _sys
        print(f"Installation de {pkg}...")
        _sp.run([_sys.executable, "-m", "pip", "install", pkg, "--quiet"], check=False)

_auto_install("openpyxl")
_auto_install("requests")
_auto_install("beautifulsoup4", "bs4")


SETTINGS_PATH = Path(__file__).parent / "settings.json"

def load_settings() -> dict:
    try:
        if SETTINGS_PATH.exists():
            return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}

def save_settings(settings: dict):
    try:
        SETTINGS_PATH.write_text(json.dumps(settings, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


BG = "#0f1117"; CARD = "#1a1d27"; ACCENT = "#6366f1"
TEXT = "#e2e8f0"; MUTED = "#94a3b8"; DIVIDER = "#252836"
GREEN = "#4ade80"; RED = "#f87171"; BLUE = "#93c5fd"
CARD2 = "#20243a"


def _load_mira_logo(size=200):
    
    here = Path(__file__).parent

    for fname in ["CTInfer_logo.png", "CTInfer.ico"]:
        src = here / fname
        if src.exists():
            try:
                from PIL import Image, ImageTk
                img = Image.open(str(src)).convert("RGBA").resize((size, size), Image.LANCZOS)
                return ImageTk.PhotoImage(img)
            except Exception:
                continue
    return None


class SplashScreen(tk.Toplevel):
    

    def __init__(self, root):
        super().__init__(root)
        self.overrideredirect(True)
        self.configure(bg="#0f1117")
        self.attributes("-topmost", True)

        w, h = 520, 440
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.configure(highlightbackground="#6366f1", highlightthickness=2)
        self.resizable(False, False)


        cv = tk.Canvas(self, width=520, height=300,
                       bg="#0f1117", highlightthickness=0)
        cv.pack()

        cx, cy = 260, 130


        import math
        for hx, hy in [(0,-1),(0,1),(1,-0.5),(1,0.5),(-1,-0.5),(-1,0.5)]:
            hcx = cx + int(hx * 68)
            hcy = cy + int(hy * 68)
            hr = 38
            pts = []
            for i in range(6):
                px = hcx + hr * math.cos(math.radians(60*i - 30))
                py = hcy + hr * math.sin(math.radians(60*i - 30))
                pts.extend([px, py])
            cv.create_polygon(pts, outline="#6366f1", fill="", width=1,
                              stipple="gray12")


        for r, w_ring in [(110, 1), (88, 1)]:
            cv.create_oval(cx-r, cy-r, cx+r, cy+r,
                           outline="#6366f1", width=w_ring, dash=(4, 6))


        bonds = [
            (cx-58, cy-44, 10, "#4f46e5"),
            (cx+58, cy-44, 10, "#4f46e5"),
            (cx-58, cy+44, 8,  "#6366f1"),
            (cx+58, cy+44, 8,  "#6366f1"),
            (cx,    cy-82, 7,  "#818cf8"),
            (cx,    cy+82, 7,  "#818cf8"),
        ]
        for bx, by, r, col in bonds:
            cv.create_line(cx, cy, bx, by, fill=col, width=2)
            cv.create_oval(bx-r, by-r, bx+r, by+r, fill=col, outline="")


        for r, col in [(30,"#1e1b4b"),(24,"#4f46e5"),(15,"#6366f1"),(8,"#a5b4fc")]:
            cv.create_oval(cx-r, cy-r, cx+r, cy+r, fill=col, outline="")


        cv.create_text(cx, 220, text="CTInfer",
                       font=("Segoe UI", 48, "bold"),
                       fill="#ffffff")


        cv.create_text(cx, 258, text="COMPOUND TARGET INFERENCE TOOL",
                       font=("Segoe UI", 9),
                       fill="#818cf8")


        cv.create_line(150, 275, 370, 275, fill="#4f46e5", width=1)


        prog_frame = tk.Frame(self, bg="#0f1117")
        prog_frame.pack(fill="x", padx=40)
        self._prog_var = tk.DoubleVar(value=0)
        style = ttk.Style()
        style.configure("mira.Horizontal.TProgressbar",
                         troughcolor="#1a1d27", background="#6366f1",
                         bordercolor="#0f1117", lightcolor="#6366f1",
                         darkcolor="#4f46e5")
        self._prog = ttk.Progressbar(prog_frame,
                                     variable=self._prog_var,
                                     style="mira.Horizontal.TProgressbar",
                                     mode="determinate", length=340)
        self._prog.pack()

        self._status = tk.Label(self, text="Initialisation...",
                                bg="#0f1117", fg="#6366f1",
                                font=("Segoe UI", 9))
        self._status.pack(pady=(8, 0))
        tk.Label(self, text="v1.0", bg="#0f1117", fg="#374151",
                 font=("Segoe UI", 8)).pack(side="bottom", pady=8)
        self.update()

    def set_status(self, msg, pct=None):
        self._status.config(text=msg)
        if pct is not None:
            self._prog_var.set(pct)
        self.update()

    def close(self):
        self._prog_var.set(100)
        self.update()
        import time as _t; _t.sleep(3.0)
        self.destroy()


class PubChemApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CTInfer — Compound Target Inference Tool")
        self.geometry("1200x820")
        self.minsize(1000, 700)
        self.configure(bg=BG)


        _ico = Path(__file__).parent / "CTInfer.ico"
        if _ico.exists():
            try:
                self.iconbitmap(str(_ico))
            except Exception:
                pass

        settings = load_settings()
        self._lang = settings.get("lang", DEFAULT_LANG)
        if self._lang not in LANG_CODES:
            self._lang = DEFAULT_LANG


        splash = SplashScreen(self)
        self.withdraw()

        splash.set_status("Chargement de l'interface...", 20)
        self._styles()
        splash.set_status("Initialisation des composants...", 50)
        self._build()
        splash.set_status("Vérification des dépendances...", 80)
        import time as _t; _t.sleep(0.3)
        splash.set_status("Prêt !", 100)
        splash.close()
        self.deiconify()

    def _build(self):

        top = tk.Frame(self, bg=BG)
        top.pack(fill="x", padx=28, pady=(18, 0))
        tk.Label(top, text="CTInfer", bg=BG, fg="#fff",
                 font=("Segoe UI", 22, "bold")).pack(side="left")
        tk.Label(top, text=" Compound Target Inference Tool", bg=BG, fg=ACCENT,
                 font=("Segoe UI", 14, "bold")).pack(side="left", pady=(6, 0))


        lang_frame = tk.Frame(top, bg=BG)
        lang_frame.pack(side="right", pady=(2, 0))

        prev_btn = tk.Label(lang_frame, text="◀", bg=BG, fg=MUTED,
                            font=("Segoe UI", 11), cursor="hand2")
        prev_btn.pack(side="left", padx=(0, 4))
        prev_btn.bind("<Button-1>", lambda e: self._change_language(-1))

        self._lang_flag_img = self._flag_image(self._lang)
        self._lang_flag_lbl = tk.Label(
            lang_frame, bg=BG, cursor="hand2",
            image=self._lang_flag_img if self._lang_flag_img else None,
            text="" if self._lang_flag_img else LANG_FLAGS.get(self._lang, "?"))
        self._lang_flag_lbl.pack(side="left", padx=4)
        self._lang_flag_lbl.bind("<Button-1>", lambda e: self._change_language(1))

        self._lang_name_lbl = tk.Label(
            lang_frame, text=LANG_NAMES.get(self._lang, "Français"),
            bg=BG, fg=MUTED, font=("Segoe UI", 9), cursor="hand2")
        self._lang_name_lbl.pack(side="left", padx=(2, 4))
        self._lang_name_lbl.bind("<Button-1>", lambda e: self._change_language(1))

        next_btn = tk.Label(lang_frame, text="▶", bg=BG, fg=MUTED,
                            font=("Segoe UI", 11), cursor="hand2")
        next_btn.pack(side="left")
        next_btn.bind("<Button-1>", lambda e: self._change_language(1))

        tk.Frame(self, bg=DIVIDER, height=1).pack(fill="x", padx=28, pady=12)


        self._tab_pubchem = tk.Frame(self, bg=BG)
        self._tab_pubchem.pack(fill="both", expand=True, padx=28, pady=(0, 18))
        self._build_pubchem_tab()

    def _change_language(self, direction: int):
        
        idx = LANG_CODES.index(self._lang) if self._lang in LANG_CODES else 0
        idx = (idx + direction) % len(LANG_CODES)
        self._lang = LANG_CODES[idx]

        settings = load_settings()
        settings["lang"] = self._lang
        save_settings(settings)

        self._rebuild_ui()

    def _rebuild_ui(self):
        
        for widget in self.winfo_children():
            widget.destroy()
        self._build()

    def _styles(self):
        s = ttk.Style(self); s.theme_use("clam")
        s.configure("TFrame",          background=BG)
        s.configure("TLabel",          background=BG, foreground=TEXT, font=("Segoe UI", 10))
        s.configure("Primary.TButton", background=ACCENT, foreground="#fff",
                    font=("Segoe UI", 10, "bold"), borderwidth=0, padding=(14, 9))
        s.configure("Ghost.TButton",   background="#1e2130", foreground=MUTED,
                    font=("Segoe UI", 10), borderwidth=0, padding=(10, 7))
        s.configure("Danger.TButton",  background="#7f1d1d", foreground="#fca5a5",
                    font=("Segoe UI", 10), borderwidth=0, padding=(10, 7))
        s.map("Primary.TButton", background=[("active","#4f46e5"),("disabled","#374151")],
              foreground=[("disabled","#6b7280")])
        s.map("Ghost.TButton",  background=[("active","#2d3148")])
        s.map("Danger.TButton", background=[("active","#991b1b")])
        s.configure("bar.Horizontal.TProgressbar",
                    troughcolor=CARD, background=ACCENT, borderwidth=0, thickness=6)
        s.configure("TNotebook",       background=BG, borderwidth=0)
        s.configure("TNotebook.Tab",   background=CARD, foreground=MUTED,
                    font=("Segoe UI", 10), padding=(16, 8))
        s.map("TNotebook.Tab",
              background=[("selected", ACCENT)],
              foreground=[("selected", "#fff")])


    def _flag_image(self, lang_code: str):
        
        import base64
        b64 = FLAG_IMAGES_B64.get(lang_code, "")
        if not b64:
            return None
        try:
            data = base64.b64decode(b64)
            return tk.PhotoImage(data=data)
        except Exception:
            return None


    def _build_pubchem_tab(self):
        import tkinter.scrolledtext as _st
        tab = self._tab_pubchem
        tab.columnconfigure(0, weight=1)
        tab.columnconfigure(1, weight=2)
        tab.rowconfigure(0, weight=1)

        def card(parent, title):
            f = tk.Frame(parent, bg=CARD); f.pack(fill="x", pady=(0,6))
            tk.Label(f, text=title, bg=CARD, fg="#fff",
                     font=("Segoe UI", 9, "bold"), padx=12, pady=7).pack(anchor="w")
            tk.Frame(f, bg=DIVIDER, height=1).pack(fill="x")
            inner = tk.Frame(f, bg=CARD, padx=12, pady=8); inner.pack(fill="x")
            return inner


        left_outer = tk.Frame(tab, bg=BG)
        left_outer.grid(row=0, column=0, sticky="nsew", padx=(0,6), pady=8)
        left_outer.rowconfigure(0, weight=1); left_outer.columnconfigure(0, weight=1)
        lcv = tk.Canvas(left_outer, bg=BG, highlightthickness=0)
        lsb = tk.Scrollbar(left_outer, orient="vertical", command=lcv.yview)
        lcv.configure(yscrollcommand=lsb.set)
        lsb.pack(side="right", fill="y")
        lcv.pack(side="left", fill="both", expand=True)
        left = tk.Frame(lcv, bg=BG)
        _lw = lcv.create_window((0,0), window=left, anchor="nw")
        left.bind("<Configure>", lambda e: lcv.configure(scrollregion=lcv.bbox("all")))
        lcv.bind("<Configure>", lambda e: lcv.itemconfig(_lw, width=e.width))
        lcv.bind("<MouseWheel>", lambda e: lcv.yview_scroll(int(-1*(e.delta/120)), "units"))


        cs = card(left, t("pc_source_title", self._lang))
        self._pc_file_var = tk.StringVar(value=t("pc_no_file", self._lang))
        tk.Label(cs, textvariable=self._pc_file_var, bg=CARD, fg=ACCENT,
                 font=("Segoe UI", 8), wraplength=200, anchor="w").pack(fill="x")
        ttk.Button(cs, text=t("pc_load_btn", self._lang),
                   command=self._pc_load_file).pack(fill="x", pady=(6,0))
        tk.Label(cs, text=t("pc_cid_col", self._lang), bg=CARD, fg=MUTED,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(8,1))
        self._pc_col_var = tk.StringVar()
        self._pc_col_cb = ttk.Combobox(cs, textvariable=self._pc_col_var,
                                        state="readonly", font=("Segoe UI", 8))
        self._pc_col_cb.pack(fill="x")
        self._pc_id_type = tk.StringVar(value="cid")


        co = card(left, t("pc_output_title", self._lang))
        self._pc_out_var = tk.StringVar(value=t("pc_no_folder", self._lang))
        tk.Label(co, textvariable=self._pc_out_var, bg=CARD, fg=ACCENT,
                 font=("Segoe UI", 8), wraplength=200, anchor="w").pack(fill="x")
        ttk.Button(co, text=t("pc_choose_dir", self._lang),
                   command=self._pc_browse_output).pack(fill="x", pady=(6,0))


        csrc = card(left, "3.  Sources de recherche")
        self._use_mce = tk.BooleanVar(value=True)
        self._use_pc  = tk.BooleanVar(value=True)
        tk.Checkbutton(csrc, text="MedChemExpress (MCE)",
                       variable=self._use_mce,
                       bg=CARD, fg=TEXT, selectcolor="#0f1117",
                       activebackground=CARD, activeforeground=TEXT,
                       font=("Segoe UI", 9)).pack(anchor="w")
        tk.Checkbutton(csrc, text="PubChem",
                       variable=self._use_pc,
                       bg=CARD, fg=TEXT, selectcolor="#0f1117",
                       activebackground=CARD, activeforeground=TEXT,
                       font=("Segoe UI", 9)).pack(anchor="w")
        self._use_cansar = tk.BooleanVar(value=True)
        tk.Checkbutton(csrc, text="CanSAR",
                       variable=self._use_cansar,
                       bg=CARD, fg=TEXT, selectcolor="#0f1117",
                       activebackground=CARD, activeforeground=TEXT,
                       font=("Segoe UI", 9)).pack(anchor="w")
        tk.Label(csrc,
                 text="  ↳ Homo Sapiens uniquement · Mean Potency < 1000 nM",
                 bg=CARD, fg=MUTED, font=("Segoe UI", 7, "italic"),
                 wraplength=200, justify="left").pack(anchor="w", pady=(0,2))


        ck = card(left, "4.  Mots-cles (optionnel)")
        tk.Label(ck, text=t("pc_terms_label", self._lang), bg=CARD, fg=MUTED,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0,2))
        self._pc_keywords_txt = tk.Text(ck, bg="#0f1117", fg=TEXT,
                                         font=("Segoe UI", 8), relief="flat",
                                         height=4, insertbackground=ACCENT)
        self._pc_keywords_txt.pack(fill="x")
        self._pc_keywords_txt.insert("1.0",
            "inhibitor\ninhibitors\ninhibit\ninhibition of\n"
            "block\nantagonist\nsuppress\ntarget\n"
            "degrader\ndegradation\nPROTAC\ndegrade\n"
            "pathway\nbinds to\nthrough inhibition")
        tk.Label(ck, text=t("pc_exclude_label", self._lang), bg=CARD, fg=MUTED,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(6,2))
        tk.Label(ck,
                 text="Si ces mots precedent un inhibiteur, la cible est ignoree. "
                      "Ex: 'not inhibits X' -> X ignore car precede de 'not'",
                 bg=CARD, fg="#6b7280", font=("Segoe UI", 7),
                 wraplength=200, justify="left").pack(anchor="w")
        self._pc_exclude_txt = tk.Text(ck, bg="#0f1117", fg=TEXT,
                                        font=("Segoe UI", 8), relief="flat",
                                        height=2, insertbackground=ACCENT)
        self._pc_exclude_txt.pack(fill="x")
        self._pc_exclude_txt.insert("1.0", "not\nno ")
        self._pc_window_var = tk.IntVar(value=8)


        right = tk.Frame(tab, bg=BG)
        right.grid(row=0, column=1, sticky="nsew", pady=8)
        right.rowconfigure(2, weight=2)
        right.rowconfigure(4, weight=1)
        right.columnconfigure(0, weight=1)


        self._pc_btn = ttk.Button(right, text=t("pc_run_btn", self._lang),
                                   style="Primary.TButton",
                                   command=self._pc_start)
        self._pc_btn.grid(row=0, column=0, sticky="ew", pady=(0,4))
        self._pc_progress_var = tk.StringVar(value=t("pc_run_hint", self._lang))
        tk.Label(right, textvariable=self._pc_progress_var, bg=BG, fg=MUTED,
                 font=("Segoe UI", 8)).grid(row=1, column=0, sticky="w", pady=(0,8))


        tk.Label(right, text=t("pc_results", self._lang), bg=BG, fg="#fff",
                 font=("Segoe UI", 9, "bold")).grid(row=2, column=0, sticky="nw")
        res_frame = tk.Frame(right, bg=CARD)
        res_frame.grid(row=2, column=0, sticky="nsew", pady=(20,8))
        res_frame.rowconfigure(0, weight=1); res_frame.columnconfigure(0, weight=1)
        self._pc_res_canvas = tk.Canvas(res_frame, bg=CARD, highlightthickness=0)
        pc_vsb = tk.Scrollbar(res_frame, orient="vertical", command=self._pc_res_canvas.yview)
        pc_hsb = tk.Scrollbar(res_frame, orient="horizontal", command=self._pc_res_canvas.xview)
        self._pc_res_canvas.configure(yscrollcommand=pc_vsb.set, xscrollcommand=pc_hsb.set)
        pc_vsb.grid(row=0, column=1, sticky="ns")
        pc_hsb.grid(row=1, column=0, sticky="ew")
        self._pc_res_canvas.grid(row=0, column=0, sticky="nsew")
        self._pc_res_inner = tk.Frame(self._pc_res_canvas, bg=CARD)
        _rw = self._pc_res_canvas.create_window((0,0), window=self._pc_res_inner, anchor="nw")
        def _update_scroll(e=None):
            self._pc_res_canvas.configure(scrollregion=self._pc_res_canvas.bbox("all"))
        self._pc_res_inner.bind("<Configure>", _update_scroll)


        tk.Label(right, text=t("pc_log", self._lang), bg=BG, fg="#fff",
                 font=("Segoe UI", 9, "bold")).grid(row=3, column=0, sticky="w")
        self._pc_log_box = _st.ScrolledText(right, bg="#0f1117", fg=TEXT,
                                             font=("Courier New", 8), relief="flat",
                                             state="disabled", wrap="word", height=7)
        self._pc_log_box.grid(row=4, column=0, sticky="nsew")
        self._pc_log_box.tag_config("ok",      foreground="#4ade80")
        self._pc_log_box.tag_config("error",   foreground="#f87171")
        self._pc_log_box.tag_config("info",    foreground="#60a5fa")
        self._pc_log_box.tag_config("warning", foreground="#fbbf24")

        self._pc_results = []


    def _pc_log(self, msg: str, level: str = "normal"):
        def _w():
            self._pc_log_box.config(state="normal")
            tag = {"ok":"ok","error":"error","info":"info","warning":"warning"}.get(level,"")
            self._pc_log_box.insert("end", msg + "\n", tag)
            n = int(self._pc_log_box.index("end-1c").split(".")[0])
            if n > 1000:
                self._pc_log_box.delete("1.0", f"{n-1000}.0")
            self._pc_log_box.see("end")
            self._pc_log_box.config(state="disabled")
        self.after(0, _w)


    def _pc_load_file(self):
        
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Charger le fichier source",
            filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("CSV", "*.csv"), ("Tous", "*.*")]
        )
        if not path: return
        try:
            p = Path(path)
            if p.suffix.lower() in (".xlsx", ".xls"):
                try:
                    import openpyxl
                except ImportError:
                    import subprocess, sys
                    messagebox.showinfo("Installation",
                        "openpyxl n'est pas installé.\n\n"
                        "Ferme l'app, ouvre un terminal et lance :\n"
                        "pip install openpyxl\n\nPuis relance l'app.")
                    return
                wb = openpyxl.load_workbook(str(p), data_only=True)
                ws = wb.active
                first_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
                headers = [str(c.value or "").strip() for c in first_row if c.value is not None]
                self._pc_data_path = str(p)
                self._pc_file_type = "excel"
            else:
                import csv
                with open(p, newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    headers = [h.strip() for h in next(reader, [])]
                self._pc_data_path = str(p)
                self._pc_file_type = "csv"

            self._pc_file_var.set(p.name)

            self._pc_col_cb["values"] = headers
            cid_col = next((h for h in headers if "cid" in h.lower()), None)
            if cid_col:
                self._pc_col_var.set(cid_col)
            elif headers:
                self._pc_col_var.set(headers[0])
        except ImportError:
            messagebox.showerror("Module manquant",
                "openpyxl n'est pas installé.\n\n"
                "Ferme l'app, lance cette commande dans un terminal :\n"
                "pip install openpyxl\n\nPuis relance l'app.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de charger :\n{e}")


    def _pc_browse_output(self):
        folder = filedialog.askdirectory(title="Dossier de sortie")
        if folder:
            self._pc_out_var.set(folder)


    def _pc_render_results(self):
        
        for w in self._pc_res_inner.winfo_children():
            w.destroy()
        if not self._pc_results:
            tk.Label(self._pc_res_inner, text="Aucun résultat", bg=CARD, fg=MUTED,
                     font=("Segoe UI", 9), padx=12, pady=8).pack()
            return
        headers = ["Composé", "Cible(s) CanSAR", "Cible(s) MCE", "Description MCE", "Cible(s) PubChem", "Description PubChem", "Refs MCE", "Refs PubChem"]
        widths  = [18, 18, 18, 35, 18, 35, 25, 25]
        hdr = tk.Frame(self._pc_res_inner, bg="#1a1d27")
        hdr.pack(fill="x")
        for h, w in zip(headers, widths):
            tk.Label(hdr, text=h, bg="#1a1d27", fg=ACCENT,
                     font=("Segoe UI", 8, "bold"), width=w,
                     anchor="w", padx=4, pady=4).pack(side="left")
        tk.Frame(self._pc_res_inner, bg=DIVIDER, height=1).pack(fill="x")
        for i, row in enumerate(self._pc_results):
            bg = "#0f1117" if i % 2 == 0 else "#13161f"
            r = tk.Frame(self._pc_res_inner, bg=bg); r.pack(fill="x")
            targets_cansar = ("; ".join(row.get("targets_cansar", [])) or "—")[:20]
            targets_mce    = ("; ".join(row.get("targets_mce", [])) or "—")[:20]
            targets_pc     = ("; ".join(row.get("targets_pc", [])) or "—")[:20]
            desc_mc = (row.get("desc_mce","") or "—")[:60]
            desc_pc = (row.get("desc_pubchem","") or "—")[:60]
            refs_mc = (row.get("refs_mce","") or "—")[:40]
            refs_pc = (row.get("refs_pubchem","") or "—")[:40]
            for val, w in zip([
                row.get("identifier",""),
                targets_cansar, targets_mce, desc_mc, targets_pc, desc_pc, refs_mc, refs_pc,
            ], widths):
                tk.Label(r, text=str(val), bg=bg, fg=TEXT,
                         font=("Segoe UI", 7), width=w,
                         anchor="w", padx=4, pady=3).pack(side="left")


    def _pc_start(self):
        
        if not hasattr(self, "_pc_data_path"):
            messagebox.showwarning("PubChem", "Charge un fichier source d'abord.")
            return
        col = self._pc_col_var.get()
        if not col:
            messagebox.showwarning("PubChem", "Selectionne la colonne des identifiants.")
            return
        output_dir = self._pc_out_var.get()
        if not output_dir or output_dir == t("pc_no_folder", self._lang):
            messagebox.showwarning("PubChem", "Choisis un dossier de sortie.")
            return

        if not self._use_mce.get() and not self._use_pc.get() and not self._use_cansar.get():
            messagebox.showwarning("CTInfer", "Coche au moins une source (MCE, PubChem ou CanSAR).")
            return

        keywords = [k.strip() for k in self._pc_keywords_txt.get("1.0","end").splitlines() if k.strip()]
        excludes = [e.strip() for e in self._pc_exclude_txt.get("1.0","end").splitlines() if e.strip()]
        window   = self._pc_window_var.get()
        id_type  = self._pc_id_type.get()
        use_mce    = self._use_mce.get()
        use_pc     = self._use_pc.get()
        use_cansar = self._use_cansar.get()

        self._pc_btn.config(state="disabled", text="Recherche en cours...")
        self._pc_results = []
        for w in self._pc_res_inner.winfo_children():
            w.destroy()

        import threading
        threading.Thread(
            target=self._pc_run,
            args=(self._pc_data_path, self._pc_file_type, col,
                  keywords, excludes, window, id_type, output_dir,
                  use_mce, use_pc, use_cansar),
            daemon=True
        ).start()


    def _pc_run(self, data_path, file_type, col, keywords, excludes,
                window, id_type, output_dir, use_mce=True, use_pc=True, use_cansar=False):
        
        import csv as _csv, re, time, json as _json
        from pathlib import Path as _Path

        try:
            import requests as _req
            from bs4 import BeautifulSoup as _BS
        except ImportError:
            self._pc_log("Module manquant : pip install requests beautifulsoup4", "error")
            self.after(0, lambda: self._pc_btn.config(state="normal", text=t("pc_run_btn", self._lang)))
            return

        SESSION = _req.Session()
        SESSION.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
        })

        JUNK_START = {"a","an","the","its","this","these","both","many","some","non",
                      "it","not","no","their","dual","potent","selective","specific",
                      "broad","novel","orally","cancer","tumor","cell","which","that",
                      "activity","potential","compound","drug","agent","treatment","new",
                      "human","natural","various","multiple","several","different"}
        JUNK_FULL  = {"inhibitor","blocker","antagonist","kinase","receptor","enzyme",
                      "suppression","inhibition","pathway","signaling","activation"}

        JUNK_PATTERNS = [
            r"^[A-Z]+\d+/[A-Z]+$",
            r"^Hep[A-Z]\d",
            r"^MCF\d",
            r"^HCC\b",
            r"^HepG\d",
        ]

        def extract_targets(compound_name, text):
            
            targets = []; kw_found = ""
            if not text: return targets, kw_found


            if text:
                first_sent = re.split(r"[.!?]\s", text)[0]
                prefix = re.escape(compound_name.split("(")[0].strip()[:10])
                m2 = re.search(
                    r"((?:[A-Z][A-Za-z0-9\-]{1,20}(?:/[A-Z][A-Za-z0-9\-]{1,20})*"
                    r"(?:\s+[A-Za-z]+){0,3}?)\s+Inhibitor)\s+" + prefix,
                    first_sent, re.IGNORECASE)
                if m2:
                    tp = re.sub(r"\s+Inhibitor$", "", m2.group(1), flags=re.IGNORECASE).strip()
                    for p in tp.split("/"):
                        p = p.strip()
                        if len(p) >= 2: targets.append(p)
                    if not kw_found: kw_found = "inhibitor (title)"


            for m in re.finditer(
                r"[A-Za-z][A-Za-z0-9\s\-]{2,50}?\s*\(([A-Z][A-Z0-9\-]{1,12})\)\s*"
                r"(?:inhibitor|inhibitors|blocker|antagonist|degrader|PROTAC)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                if len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "inhibitor of"


            for m in re.finditer(
                r"(?:EC\s+[\d.]+|[A-Za-z0-9\-]{2,20})\s*\(([a-zA-Z][a-zA-Z0-9\s\-]{2,60}?)\)\s*"
                r"(?:inhibitor|inhibitors|blocker|antagonist|degrader|PROTAC)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                if (len(tgt) >= 2 and not tgt.isupper()
                        and not re.match(r"^[A-Z][A-Z0-9\-]{1,12}$", tgt)):
                    targets.append(tgt)
                    if not kw_found: kw_found = "inhibitor of (EC)"


            for m in re.finditer(
                r"(?:inhibitor|inhibitors)\s+of\s+(?:the\s+)?(.{3,80}?)(?=\s*[,\.]\s|\s+class\b|\s+with\b|\s+which\b|\s+and\b|$)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                acro = re.search(r"\(([A-Z][A-Z0-9\-]{1,12})\)", tgt)
                if acro:
                    targets.append(acro.group(1))
                else:
                    clean = re.sub(r"\s*\([^)]+\)", "", tgt).strip().rstrip(".,;:")
                    if clean and len(clean) < 50 and re.search(r"[A-Z]", clean) and not re.search(r"\b(is|are|was)\b", clean):
                        targets.append(clean)
                if not kw_found: kw_found = "inhibitor of"


            for m in re.finditer(
                r"\b([A-Z][A-Z0-9\-]{1,20}(?:/[A-Z][A-Z0-9\-]{1,20})*"
                r"(?:\s+(?:class\s+[IVX]+|kinase|receptor|deacetylase))?)"
                r"\s+(?:inhibitor|inhibitors|blocker|antagonist|degrader|PROTAC)\b",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                if len(tgt) >= 2 and re.search(r"[A-Z]{2,}", tgt):
                    targets.append(tgt)
                    if not kw_found: kw_found = "inhibitor (suffix)"


            for m in re.finditer(
                r"\b([A-Z][A-Za-z]*(?:-[A-Za-z]+)?(?:\s+[A-Za-z0-9]{1,3})?)\s+"
                r"(?:kinase|receptor|phosphatase|deacetylase|protease)\s+"
                r"(?:inhibitor|inhibitors|blocker|antagonist|degrader|PROTAC)\b",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                if len(tgt) >= 2 and not re.match(r"^(?:An?|The|This|It|That)$", tgt, re.IGNORECASE):
                    targets.append(tgt)
                    if not kw_found: kw_found = "inhibitor (named kinase/receptor)"


            for m in re.finditer(
                r"(?:dual|selective|potent|specific)\s+([\w\-]+)\s+and\s+([\w\-]+)\s+inhibitor",
                text, re.IGNORECASE):
                for g in [m.group(1), m.group(2)]:
                    if g and len(g) >= 2 and re.search(r"[A-Z]", g):
                        targets.append(g.strip())
                if not kw_found: kw_found = "dual inhibitor"


            for m in re.finditer(
                r"inhibits?\s+(?:the\s+)?([A-Z][A-Z0-9\+\-]{1,20}"
                r"(?:/[A-Z][A-Z0-9\-]{1,20})?"
                r"(?:\s+(?:kinase|receptor|deacetylase|ATPase|synthase))?)",
                text):
                tgt = m.group(1).strip()
                if len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "inhibits"


            for m in re.finditer(
                r"(?:designed|used|developed)\s+to\s+inhibit\s+(?:the\s+)?(.{4,80}?)(?=[,\.]|$)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                acro = re.search(r"\(([A-Z][A-Z0-9\-]{1,12})\)", tgt)
                if acro:
                    targets.append(acro.group(1))
                else:
                    clean = re.sub(r"\s*\([^)]+\)", "", tgt).strip()
                    if clean and len(clean) < 50 and re.search(r"[A-Z]", clean):
                        targets.append(clean)
                if not kw_found: kw_found = "designed to inhibit"


            for m in re.finditer(
                r"(?:degrader|PROTAC)\s+of\s+(?:the\s+)?(.{3,60}?)(?=[,\.\s]|$)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip().rstrip(".,;:")
                acro = re.search(r"\(([A-Z][A-Z0-9\-]{1,12})\)", tgt)
                if acro: tgt = acro.group(1)
                else: tgt = re.sub(r"\s*\([^)]+\)", "", tgt).strip()
                if tgt and len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "degrader/PROTAC"

            for m in re.finditer(
                r"degrades?\s+(?:the\s+)?([A-Z][A-Za-z0-9\-]{1,30})",
                text, re.IGNORECASE):
                tgt = m.group(1).strip()
                if len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "degrades"

            for m in re.finditer(
                r"degradation\s+of\s+(?:the\s+)?(.{3,60}?)(?=[,\.\s]|$)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip().rstrip(".,;:")
                acro = re.search(r"\(([A-Z][A-Z0-9\-]{1,12})\)", tgt)
                if acro: tgt = acro.group(1)
                else: tgt = re.sub(r"\s*\([^)]+\)", "", tgt).strip()
                if tgt and len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "degradation of"


            for m in re.finditer(
                r"blocks?\s+(?:the\s+)?([A-Z][A-Za-z0-9\-\s]{2,30}?)"
                r"(?=\s*[\(\,\.]|\s+and\b|$)", text):
                tgt = m.group(1).strip()
                if re.search(r"[A-Z]{2,}|kinase|receptor", tgt, re.IGNORECASE):
                    targets.append(tgt)
                    if not kw_found: kw_found = "blocks"


            for m in re.finditer(
                r"(?:targets?|binds?\s+to)\s+(?:the\s+)?([A-Z][A-Za-z0-9\-\s]{2,30}?)"
                r"(?=\s*[\(\,\.]|\s+and\b|$)", text, re.IGNORECASE):
                tgt = m.group(1).strip()
                if re.search(r"[A-Z]{2,}|kinase|receptor", tgt, re.IGNORECASE):
                    targets.append(tgt)
                    if not kw_found: kw_found = "targets/binds"


            for m in re.finditer(
                r"inhibition\s+of\s+(?:the\s+)?([A-Za-z0-9][A-Za-z0-9\-/]{1,30}"
                r"(?:\s+(?:pathway|signaling|axis|kinase|receptor))?)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip().rstrip(".,;:")
                if len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "inhibition of"


            for m in re.finditer(
                r"suppression\s+of\s+(?:the\s+)?([A-Za-z0-9][A-Za-z0-9\-/]{1,30}"
                r"(?:\s+(?:pathway|signaling|axis))?)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip().rstrip(".,;:")
                if len(tgt) >= 2:
                    targets.append(tgt)
                    if not kw_found: kw_found = "suppression of"


            for m in re.finditer(
                r"([A-Z][A-Z0-9]{1,8}(?:/[A-Za-z][A-Za-z0-9]{1,8}){1,5})"
                r"(?:\s+(?:pathway|signaling|axis|inhibitor|kinase|cascade))?",
                text):
                tgt = m.group(1).strip()
                if len(tgt) >= 3 and "/" in tgt:
                    targets.append(tgt)
                    if not kw_found: kw_found = "pathway/axis"


            for m in re.finditer(
                r"through\s+(?:the\s+)?(?:inhibition\s+of\s+)?([A-Za-z0-9][A-Za-z0-9\-/]{2,30}"
                r"(?:\s+(?:pathway|signaling|axis|kinase|receptor))?)",
                text, re.IGNORECASE):
                tgt = m.group(1).strip().rstrip(".,;:")
                if len(tgt) >= 3 and re.search(r"[A-Z]", tgt):
                    targets.append(tgt)
                    if not kw_found: kw_found = "through"


            clean = []
            for tgt in targets:
                tgt = tgt.strip().rstrip(".,;:) ")
                tgt = re.sub(r"\s+", " ", tgt)
                if not tgt or len(tgt) < 2 or len(tgt) > 60: continue
                if tgt.split()[0].lower() in JUNK_START: continue
                if tgt.lower() in JUNK_FULL: continue
                if not re.search(r"[A-Za-z]", tgt): continue
                if re.search(r"\b(is|are|was|were|has|have|can|will|would)\b", tgt, re.IGNORECASE): continue
                if any(re.search(pat, tgt) for pat in JUNK_PATTERNS): continue
                if tgt not in clean: clean.append(tgt)

            seen = set(); final = []
            for tgt in sorted(clean, key=len):
                t_low = re.sub(r"[-\s]*(alpha|beta|gamma)$", "", tgt.lower()).strip()
                if not any(t_low in s or s in t_low for s in seen):
                    seen.add(t_low); seen.add(tgt.lower()); final.append(tgt)

            return final[:6], kw_found

        def fetch_mce_full(name, pw_page=None):
            
            desc = ""; refs = []; mce_targets = []
            try:
                slug = re.sub(r"[^a-z0-9\-]", "-", name.lower()).strip("-")
                slug = re.sub(r"-+", "-", slug)
                product_url = f"https://www.medchemexpress.com/{slug}.html"

                html_content = ""
                page = pw_page

                try:
                    page.goto(product_url, timeout=30000,
                              wait_until="domcontentloaded")
                    try:
                        page.wait_for_selector("#product_syn", timeout=8000)
                    except Exception:
                        pass
                    html_content = page.content()


                    if "product_syn" not in html_content:
                        from urllib.parse import quote as _q
                        search_url = f"https://www.medchemexpress.com/search.html?q={_q(name)}"
                        page.goto(search_url, timeout=30000,
                                  wait_until="domcontentloaded")
                        try:
                            first = page.locator(
                                "a.product-name, .product-item a, "
                                "h3.product-title a, .search-result a"
                            ).first
                            href = first.get_attribute("href", timeout=5000)
                            if href:
                                if not href.startswith("http"):
                                    href = "https://www.medchemexpress.com" + href
                                page.goto(href, timeout=30000,
                                          wait_until="domcontentloaded")
                                try:
                                    page.wait_for_selector("#product_syn", timeout=8000)
                                except Exception:
                                    pass
                                html_content = page.content()
                        except Exception:
                            pass
                except Exception:
                    pass

                if not html_content:
                    return desc, refs, mce_targets

                soup2 = _BS(html_content, "html.parser")


                prod_syn = soup2.select_one("#product_syn")
                if prod_syn:
                    desc = prod_syn.get_text(" ", strip=True)[:3000]


                if not desc:
                    compound_word = name.split()[0]
                    for p in soup2.select("p"):
                        tgt = p.get_text(" ", strip=True)
                        if (len(tgt) > 60 and
                            compound_word.lower() in tgt.lower() and
                            not re.search(r"EUR\s*\d|En stock|recommande|vendons",
                                          tgt, re.IGNORECASE)):
                            desc = tgt[:3000]; break


                desc_el = soup2.select_one("#product_syn") or soup2
                for a_tag in desc_el.select("a[href*='/Targets/'], a[href*='/targets/']"):
                    target_name = a_tag.get_text(" ", strip=True)

                    target_name = re.sub(
                        r"^(de la|d'|des|la|le|les|l'|the|of the|of|du|de)\s+",
                        "", target_name, flags=re.IGNORECASE).strip()
                    target_name = target_name.strip(".,;:()")

                    if (target_name and 2 <= len(target_name) <= 50 and
                        not re.search(
                            r"voir|show|all|isoform|produit|product|click|cancer|"
                            r"^d'autres$|^other$|^autres$|^du cancer$|"
                            r"^apoptose$|^apoptosis$|^autophagy$|^autophagie$|"
                            r"^caspases?$|^PARP$",
                            target_name, re.IGNORECASE)):
                        mce_targets.append(target_name)


                seen_t = set(); mce_targets_clean = []
                for tgt in mce_targets:
                    if tgt.lower() not in seen_t:
                        seen_t.add(tgt.lower())
                        mce_targets_clean.append(tgt)
                mce_targets = mce_targets_clean[:10]


                for ref_sel in [
                    "#cpd_References1", "#cpd_References2",
                    "#pro_info_5", ".pro_info_ref",
                    "[id^=cpd_References]",
                ]:
                    container = soup2.select_one(ref_sel)
                    if container:

                        for p_tag in container.select("p"):
                            tgt = p_tag.get_text(" ", strip=True)
                            if re.match(r"^\[\d+\]", tgt) and len(tgt) > 20:
                                refs.append(tgt)
                        if refs: break


                if not refs:
                    for p_tag in soup2.select("p"):
                        tgt = p_tag.get_text(" ", strip=True)
                        if re.match(r"^\[\d+\]\.", tgt) and len(tgt) > 20:
                            refs.append(tgt)


                if not refs:
                    ref_blocks = re.findall(
                        r"\[\d+\]\..+?(?=\[\d+\]\.|</ul>|</li>|$)",
                        html_content, re.DOTALL)
                    for block in ref_blocks[:10]:
                        clean_ref = re.sub(r"<[^>]+>", " ", block)
                        clean_ref = re.sub(r"\s+", " ", clean_ref).strip()
                        if len(clean_ref) > 20:
                            refs.append(clean_ref)


                seen_refs = set()
                refs_clean = []
                for r_item in refs:
                    r_norm = re.sub(r"\s+", " ", r_item).strip()
                    if r_norm not in seen_refs:
                        seen_refs.add(r_norm)
                        refs_clean.append(r_norm)
                refs = refs_clean[:10]

            except Exception as e:
                self._pc_log(f"  MCE erreur : {e}", "warning")
            return desc, refs, mce_targets

        def fetch_pubchem_full(cid):
            
            desc = ""; refs = []
            try:

                r_desc = SESSION.get(
                    f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/description/JSON",
                    timeout=15)
                if r_desc.status_code == 200:
                    for info in r_desc.json().get("InformationList", {}).get("Information", []):
                        d = info.get("Description", "")
                        if d and len(d) > 20:
                            desc += d + " "
                    desc = desc.strip()[:2000]
                time.sleep(0.3)


                r = SESSION.get(
                    f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON",
                    timeout=25)
                if r.status_code == 200:
                    data = r.json()
                    texts = []
                    SKIP_SECTIONS = {
                        "canonical smiles", "isomeric smiles", "inchi", "inchikey",
                        "molecular formula", "molecular weight", "iupac name",
                        "2d structure", "3d structure", "chemical structure", "structure",
                        "names and identifiers", "computed descriptors",
                        "topological polar surface area", "xlogp3",
                        "hydrogen bond", "rotatable bond", "exact mass", "monoisotopic mass",

                        "chemical neighbors", "related chemicals",
                        "chemicalneighbor", "patent", "patents",
                        "depositor-supplied", "biological test results",
                        "literature", "associated disorders",
                        "classification", "drug categories",
                    }

                    JUNK_STRINGS = [
                        "view in pubchem search",
                        "chemicalneighbor",
                        "chemicalgenesy",
                        "chemicaldiseasen",
                        "chemicalorganism",
                        "patentchemical",
                        "patentchemic",
                        "https://pubchem.ncbi.nlm.nih.gov/gene/",
                        "https://pubchem.ncbi.nlm.nih.gov/taxonomy/",
                        "https://meshb.nlm.nih.gov/",
                        "https://patentscope.wipo.int/",
                        "link to all deposited",
                        "patents are available",
                        "gene/protein/enzyme",
                        "(related)",
                    ]

                    def _is_struct(txt):
                        if txt.startswith("InChI="): return True
                        if re.match(r'^[A-Z]{14}-[A-Z]{10}-[A-Z]$', txt): return True
                        if re.match(r'^[A-Z][a-z]?(\d+)?([A-Z][a-z]?(\d+)?)+$', txt) and len(txt) < 30: return True
                        if re.match(r'^[A-Za-z0-9@+\-\[\]()/#%=.]{10,}$', txt) and any(c in txt for c in ['[','@@','#']): return True
                        return False

                    def _ex(obj, _sec=""):
                        if isinstance(obj, dict):
                            sec = obj.get("TOCHeading", obj.get("Name", "")).lower()
                            if sec and any(s in sec for s in SKIP_SECTIONS):
                                return
                            if "StringWithMarkup" in obj:
                                for s in obj["StringWithMarkup"]:
                                    txt = s.get("String", "")
                                    tl = txt.lower()
                                    if (len(txt) > 15 and not _is_struct(txt)
                                            and not any(j in tl for j in JUNK_STRINGS)):
                                        texts.append(txt)
                            if obj.get("ReferenceNumber") and obj.get("SourceName"):
                                ref_text = f"{obj.get('SourceName','')}. {obj.get('Title','')}"
                                if obj.get("URL"):
                                    ref_text += f". {obj.get('URL','')}"
                                if len(ref_text) > 10 and ref_text not in refs:
                                    refs.append(ref_text)
                            for v in obj.values(): _ex(v)
                        elif isinstance(obj, list):
                            [_ex(i) for i in obj]
                    _ex(data)

                    extra = " ".join(texts)
                    if not desc:
                        desc = extra[:2000]
                    elif len(desc) < 300 and extra:
                        desc = (desc + " " + extra)[:2000]
                time.sleep(0.3)


                if not desc or len(desc) < 80:
                    r3 = SESSION.get(
                        f"https://pubchem.ncbi.nlm.nih.gov/compound/{cid}",
                        timeout=20)
                    if r3.status_code == 200:
                        soup = _BS(r3.text, "html.parser")

                        pharma_text = ""
                        for section in soup.find_all(["h2","h3","h4","div","p"]):
                            txt = section.get_text(" ", strip=True)
                            if "pharmacology" in txt.lower() or "biochemistry" in txt.lower():

                                nxt = section.find_next("p")
                                if nxt:
                                    pharma_text = nxt.get_text(" ", strip=True)[:500]
                                break

                        if not pharma_text:
                            for el in soup.select("p"):
                                txt_el = el.get_text(" ", strip=True)
                                if len(txt_el) > 50 and not txt_el.startswith("http"):
                                    pharma_text = txt_el[:400]
                                    break
                        if pharma_text:
                            desc = (pharma_text + " " + desc).strip()[:2000]
                    time.sleep(0.3)

            except Exception as e:
                self._pc_log(f"  PubChem erreur : {e}", "warning")
            return desc.strip(), refs[:10]


        def fetch_cansar(compound_name, pw_page=None):
            
            if pw_page is None:
                return []
            page = pw_page
            targets_cansar = []
            try:
                import re as _re2


                page.goto("https://cansar.ai/", timeout=30000, wait_until="domcontentloaded")
                page.wait_for_timeout(1500)


                search_input = page.locator("input[type='text'], input[type='search']").first
                search_input.click()
                search_input.fill(compound_name)
                page.wait_for_timeout(1500)


                try:
                    suggestion = page.locator(
                        f"text='{compound_name}'"
                    ).first
                    suggestion.click(timeout=5000)
                except Exception:

                    search_input.press("Enter")
                page.wait_for_timeout(3000)


                cansar_id = None
                cards = page.locator("text='COMPOUND'")
                n_cards = cards.count()


                all_hrefs = page.eval_on_selector_all(
                    "a[href*='/compound/']",
                    "els => els.map(e => e.getAttribute('href'))"
                )

                cname_lower = compound_name.lower().strip()
                for href in all_hrefs:
                    m_id = _re2.search(r"/compound/([0-9]+)", href or "")
                    if m_id:
                        cansar_id = m_id.group(1)
                        break

                if not cansar_id:
                    self._pc_log(f"    CanSAR: composé non trouvé", "warning")
                    return []


                ta_url = f"https://cansar.ai/compound/{cansar_id}/target-affinity"
                page.goto(ta_url, timeout=30000, wait_until="networkidle")
                page.wait_for_timeout(3000)


                try:
                    page.select_option("select#tablefilter", "Homo sapiens")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass

                try:
                    page.select_option("select#target-afinity-pagination", "50")
                    page.wait_for_timeout(600)
                except Exception:
                    pass

                rows = page.locator("table tbody tr")
                count = rows.count()
                for i in range(count):
                    row = rows.nth(i)
                    cells = row.locator("td")
                    if cells.count() < 4:
                        continue
                    try:
                        protein = cells.nth(0).inner_text().strip()
                        organism = cells.nth(2).inner_text().strip()
                        potency_txt = cells.nth(3).inner_text().strip()
                        if "homo sapiens" not in organism.lower():
                            continue
                        pt = potency_txt.lower().replace(",", ".")
                        nm_val = None
                        m2 = _re2.search(r"([\d.]+)\s*nm", pt)
                        if m2:
                            nm_val = float(m2.group(1))
                        else:
                            m3 = _re2.search(r"([\d.]+)\s*[µu]m", pt)
                            if m3:
                                nm_val = float(m3.group(1)) * 1000
                        if nm_val is not None and nm_val < 1000 and protein:
                            targets_cansar.append(protein)
                    except Exception:
                        continue

            except Exception as e:
                self._pc_log(f"  CanSAR erreur : {e}", "warning")

            return list(dict.fromkeys(targets_cansar))

        def best_sentence(text, compound_name):
            
            PHARMA_KW = [
                "inhibit", "block", "target", "suppress", "antagonist",
                "bind", "pathway", "signaling", "apoptosis", "antitumor",
                "anticancer", "antineoplastic", "proliferation", "neoplasm",
                "degrader", "degradation", "protac", "degrade",
                "activat", "modulat", "antifungal", "antibacterial",
                "antiviral", "cytotoxic", "antiproliferat",
                "potent", "selective", "orally active", "ic50", " nm",
                "agonist", "allosteric", "epigenetic", "histone",
                "bromodomain", "acetylation", "methylation", "chromatin",
                "enzyme", "protein", "cancer", "tumor", "leukemia",
            ]
            if text:
                tl = text.lower()

                if any(kw in tl for kw in PHARMA_KW):

                    sentences = re.split(r"(?<=[.!?])\s+", text)

                    for s in sentences:
                        if any(kw in s.lower() for kw in PHARMA_KW) and 15 < len(s) < 800:
                            return s.strip()

                    if len(text) < 400:
                        return text.strip()

                    for kw in PHARMA_KW:
                        idx = tl.find(kw)
                        if idx >= 0:
                            start = max(0, text.rfind(" ", 0, idx) - 50)
                            end = text.find(".", idx)
                            if end < 0 or end - start > 400: end = idx + 200
                            return text[start:end].strip()
            return ""


        identifiers = []
        try:
            if file_type == "excel":
                import openpyxl as _opx
                wb = _opx.load_workbook(data_path, data_only=True)
                ws_xl = wb.active
                headers_xl = [str(c.value or "").strip()
                               for c in list(ws_xl.iter_rows(min_row=1, max_row=1))[0]]
                col_idx = headers_xl.index(col)
                name_col_idx = next((i for i, h in enumerate(headers_xl)
                                     if "compound" in h.lower()), 0)
                for row in ws_xl.iter_rows(min_row=2, values_only=True):
                    if col_idx < len(row) and row[col_idx]:
                        identifiers.append((
                            str(row[name_col_idx]).strip() if name_col_idx < len(row) else "",
                            str(row[col_idx]).strip()
                        ))
            else:
                with open(data_path, newline="", encoding="utf-8-sig") as f:
                    reader = _csv.DictReader(f)
                    hdrs = reader.fieldnames or []
                    name_col_f = next((h for h in hdrs if "compound" in h.lower()),
                                      hdrs[0] if hdrs else col)
                    for row in reader:
                        if col in row and row[col]:
                            identifiers.append((
                                str(row.get(name_col_f, row[col])).strip(),
                                str(row[col]).strip()
                            ))
        except Exception as e:
            self._pc_log(f"Erreur lecture fichier : {e}", "error")
            self.after(0, lambda: self._pc_btn.config(
                state="normal", text=t("pc_run_btn", self._lang)))
            return

        self._pc_log(f"  {len(identifiers)} composé(s) à traiter", "info")
        results = []
        total = len(identifiers)


        _pw_instance = None
        _pw_browser  = None
        _pw_ctx      = None
        _pw_page_mce    = None
        _pw_page_cansar = None
        if use_mce or use_cansar:
            try:
                from playwright.sync_api import sync_playwright as _spw_init
                _pw_instance = _spw_init().__enter__()
                _pw_browser  = _pw_instance.chromium.launch(headless=True)
                if use_mce:
                    _pw_ctx = _pw_browser.new_context(
                        locale="en-US",
                        extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
                    )
                    _pw_ctx.add_cookies([{
                        "name": "language", "value": "en",
                        "domain": "www.medchemexpress.com", "path": "/"
                    }])
                    _pw_page_mce = _pw_ctx.new_page()
                if use_cansar:
                    _pw_page_cansar = _pw_browser.new_page(
                        extra_http_headers={"Accept-Language": "en-US,en;q=0.9"}
                    )
            except Exception as _be:
                self._pc_log(f"  Playwright init erreur : {_be}", "error")

        try:
          for idx, (name, cid_raw) in enumerate(identifiers):
            cid = re.sub(r"\.0$", "", cid_raw.strip())
            self._pc_log(f"[{idx+1}/{total}] {name} (CID: {cid})")
            self.after(0, lambda p=int((idx+1)/total*100), n=name:
                self._pc_progress_var.set(f"{p}% — {n}"))


            if use_mce and _pw_page_mce:
                mce_desc, mce_refs, mce_targets_direct = fetch_mce_full(name, pw_page=_pw_page_mce)
            else:
                mce_desc, mce_refs, mce_targets_direct = "", [], []


            if use_pc:
                pc_desc, pc_refs = fetch_pubchem_full(cid)
            else:
                pc_desc, pc_refs = "", []


            targets_mce_regex, _kw_mce_regex = extract_targets(name, mce_desc)
            targets_pc,        _kw_pc        = extract_targets(name, pc_desc)


            all_targets_mce = list(mce_targets_direct)
            for tgt in targets_mce_regex:
                if not any(tgt.lower() in x.lower() or x.lower() in tgt.lower()
                           for x in all_targets_mce):
                    all_targets_mce.append(tgt)


            all_targets_pc = list(targets_pc)

            target_mce_str = "; ".join(all_targets_mce[:8])
            target_pc_str  = "; ".join(all_targets_pc[:8])

            status = target_mce_str or target_pc_str or "aucune cible"
            self._pc_log(f"  {name} → {status[:80]}")


            mce_sentence = best_sentence(mce_desc, name)
            pc_sentence  = best_sentence(pc_desc,  name)


            mce_refs_str = "\n".join(mce_refs[:5]) if mce_refs else ""
            pc_refs_str  = "\n".join(pc_refs[:5])  if pc_refs  else ""


            cansar_tgts = []
            if use_cansar and _pw_page_cansar:
                self._pc_log(f"  CanSAR: {name}", "info")
                cansar_tgts = fetch_cansar(name, pw_page=_pw_page_cansar)
                if cansar_tgts:
                    self._pc_log(f"    CanSAR cibles: {', '.join(cansar_tgts)}", "ok")

            result = {
                "identifier":      name,
                "cid":             cid,
                "targets_mce":     all_targets_mce,
                "targets_pc":      all_targets_pc,
                "targets_cansar":  cansar_tgts,
                "desc_mce":        mce_sentence,
                "desc_pubchem":    pc_sentence,
                "refs_mce":        mce_refs_str,
                "refs_pubchem":    pc_refs_str,
            }
            results.append(result)
            self._pc_results = results
            self.after(0, self._pc_render_results)

        finally:

            try:
                if _pw_page_mce:    _pw_page_mce.close()
                if _pw_page_cansar: _pw_page_cansar.close()
                if _pw_ctx:         _pw_ctx.close()
                if _pw_browser:     _pw_browser.close()
                if _pw_instance:    _pw_instance.__exit__(None, None, None)
            except Exception:
                pass


        out_path = _Path(output_dir)
        try:
            import openpyxl as _opx2
            from openpyxl.styles import Font as _Ft, PatternFill as _PF, Alignment as _Al
            from openpyxl.utils import get_column_letter as _gcl

            wb2 = _opx2.Workbook()
            ws_r = wb2.active; ws_r.title = "Drug Targets"
            HDR  = _Ft(bold=True, color="FFFFFF", name="Arial", size=10)
            FILL_H  = _PF("solid", start_color="1F4E79")
            FILL_OK = _PF("solid", start_color="E8F5EE")
            FILL_NO = _PF("solid", start_color="FFF3CD")
            WRAP = _Al(wrap_text=True, vertical="top")
            CTR  = _Al(horizontal="center", vertical="top")

            wb2 = _opx2.Workbook()
            ws_r = wb2.active; ws_r.title = "Drug Targets"
            HDR  = _Ft(bold=True, color="FFFFFF", name="Arial", size=10)
            FILL_H  = _PF("solid", start_color="1F4E79")
            FILL_OK = _PF("solid", start_color="E8F5EE")
            FILL_NO = _PF("solid", start_color="FFF3CD")
            WRAP = _Al(wrap_text=True, vertical="top")
            CTR  = _Al(horizontal="center", vertical="top")


            col_defs = [("Compound Name", 25, lambda r: r["identifier"])]
            if use_cansar:
                col_defs.append(("Target(s) CanSAR", 30, lambda r: "; ".join(r.get("targets_cansar", []))))
            if use_mce:
                col_defs.append(("Target(s) MCE", 30, lambda r: "; ".join(r.get("targets_mce", []))))
                col_defs.append(("Description MCE", 50, lambda r: r.get("desc_mce", "")))
            if use_pc:
                col_defs.append(("Target(s) PubChem", 30, lambda r: "; ".join(r.get("targets_pc", []))))
                col_defs.append(("Description PubChem", 50, lambda r: r.get("desc_pubchem", "")))
            if use_mce:
                col_defs.append(("References MCE", 45, lambda r: r.get("refs_mce", "")))
            if use_pc:
                col_defs.append(("References PubChem", 45, lambda r: r.get("refs_pubchem", "")))

            for c, (h, w, _fn) in enumerate(col_defs, 1):
                cell = ws_r.cell(1, c, h)
                cell.font = HDR; cell.fill = FILL_H; cell.alignment = CTR
                ws_r.column_dimensions[_gcl(c)].width = w
            ws_r.row_dimensions[1].height = 22
            ws_r.freeze_panes = "A2"

            for r2, res in enumerate(results, 2):
                has = bool(res.get("targets_mce") or res.get("targets_pc") or res.get("targets_cansar"))
                for c, (_h, _w, fn) in enumerate(col_defs, 1):
                    val = fn(res)
                    cell = ws_r.cell(r2, c, val or "")
                    cell.font = _Ft(name="Arial", size=9)
                    cell.fill = FILL_OK if has else FILL_NO
                    cell.alignment = WRAP
                ws_r.row_dimensions[r2].height = 60

            last_col_letter = _gcl(len(col_defs))
            ws_r.auto_filter.ref = f"A1:{last_col_letter}{len(results)+1}"
            xl_path = out_path / "Results_targets.xlsx"
            if xl_path.exists():
                _counter = 1
                while True:
                    candidate = out_path / f"Results_targets_{_counter}.xlsx"
                    if not candidate.exists():
                        xl_path = candidate
                        break
                    _counter += 1
            wb2.save(xl_path)
            self._pc_log(f"Excel : {xl_path.name}", "ok")
        except Exception as e:
            self._pc_log(f"Erreur Excel : {e}", "error")

        n_ok = sum(1 for r in results
                   if r.get("targets_mce") or r.get("targets_pc") or r.get("targets_cansar"))
        self.after(0, lambda: self._pc_progress_var.set(
            f"Terminé — {n_ok}/{len(results)} avec cible trouvée"))
        self.after(0, lambda: self._pc_btn.config(
            state="normal", text=t("pc_run_btn", self._lang)))


if __name__ == "__main__":
    app = PubChemApp()
    app.mainloop()
